"""
원료 재고 소진 예측 & 발주 알림 앱 (Streamlit 판)

React + FastAPI 구조(src/, server/)를 단일 파이썬 프로세스로 옮긴 것이다.
백엔드가 앱 안으로 들어왔으므로 HTTP 계층과 CORS가 사라지고, OPENAI_API_KEY는
처음부터 서버 프로세스에만 존재한다.

실행:
    streamlit run app.py

지켜야 할 불변식 (PRD §8.3 — 스타일 취향이 아니라 실제로 버그를 냈던 지점들이다):
  INV-1  입고 집계는 status == 'scheduled'인 일정만. completed는 이미 재고에 반영됐다.
  INV-2  날짜는 로컬 달력 기준 YYYY-MM-DD 문자열. UTC 변환(utcnow) 금지.
  INV-3  시뮬레이션 상한 365일. 도달 못 하면 None.
  INV-4  숫자 입력이 조용히 0으로 저장되지 않게 한다. (Streamlit 적응 방식은 아래 주석 참고)
  AI-02  D-Day·발주일은 LLM이 계산하지 않는다. 앱이 계산한 값을 문장에 넣어 전달만 시킨다.
"""

from __future__ import annotations

import hashlib
import io
import json
import logging
import os
import re
import threading
from dataclasses import asdict, dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Literal
from uuid import uuid4

import altair as alt
import numpy as np
import pandas as pd
import streamlit as st

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s: %(message)s")
logger = logging.getLogger("rms")

# ===========================================================================
# 설정
# ===========================================================================

APP_ROOT = Path(__file__).resolve().parent
DATA_DIR = APP_ROOT / "data"
STORE_PATH = DATA_DIR / "store.json"
EMBED_CACHE_PATH = DATA_DIR / "embed_cache.npz"

EMBEDDING_DIM = 3072  # text-embedding-3-large

# PRD §10.3 검색 파라미터
RAG_CANDIDATE_K = 20
RAG_TOP_K = 5
RAG_MIN_SCORE = 0.30
RAG_MMR_LAMBDA = 0.5
RAG_MAX_CONTEXT_TOKENS = 1200
RAG_EMBED_BATCH = 100
EMBED_CACHE_MAX = 5000

UNIT_OPTIONS = ["kg", "g", "L", "mL", "EA", "BOX", "ton"]

STATUS_LABELS = {"safe": "안전", "warning": "발주 임박", "overdue": "발주 지연"}
STATUS_ICONS = {"safe": "🟢", "warning": "🟡", "overdue": "🔴"}
STATUS_ORDER = {"overdue": 0, "warning": 1, "safe": 2}
SCHEDULE_STATUS_LABELS = {
    "scheduled": "예정",
    "completed": "입고 완료",
    "cancelled": "취소됨",
}
LOG_TYPE_LABELS = {"incoming": "입고", "adjustment": "조정", "consumption": "사용"}

# 차트 색 — Streamlit 라이트/다크 양쪽에서 대비가 나오는 값으로 고정
COLOR_LINE = "#3d5afe"
COLOR_SAFETY = "#d97706"
COLOR_DEPLETION = "#dc2626"
COLOR_INCOMING = "#16a34a"


def _load_env() -> None:
    """.env를 읽는다. 키는 이 프로세스 밖으로 나가지 않는다."""
    try:
        from dotenv import load_dotenv

        load_dotenv(APP_ROOT / ".env")
    except ImportError:
        logger.warning("python-dotenv가 없어 .env를 읽지 못했습니다. 환경변수를 직접 설정하세요.")


_load_env()


def get_setting(key: str, default: str = "") -> str:
    """환경변수 → st.secrets 순으로 찾는다. (Streamlit Cloud 배포 대비)"""
    value = os.getenv(key)
    if value:
        return value
    try:
        secret = st.secrets.get(key)  # type: ignore[attr-defined]
        if secret:
            return str(secret)
    except Exception:  # noqa: BLE001 - secrets.toml이 없으면 예외가 난다
        pass
    return default


def get_api_key() -> str:
    key = get_setting("OPENAI_API_KEY")
    return "" if key.startswith("sk-...") else key


def get_chat_model() -> str:
    return get_setting("OPENAI_CHAT_MODEL", "gpt-5.6-sol")


def get_embedding_model() -> str:
    return get_setting("OPENAI_EMBEDDING_MODEL", "text-embedding-3-large")


# ===========================================================================
# 날짜 유틸 (INV-2)
# ===========================================================================
#
# 파이썬에서는 date.today()가 로컬 달력 날짜를 주고 isoformat()이 YYYY-MM-DD를 주므로
# 자바스크립트의 toISOString() 함정이 없다. 다만 datetime.utcnow().date()를 쓰면
# KST에서 오전 9시 이전이 전날로 밀리므로 절대 쓰지 않는다.

DATE_PATTERN = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def today_str() -> str:
    return date.today().isoformat()


def parse_date(value: str) -> date:
    if not DATE_PATTERN.match(value):
        raise ValueError(f"잘못된 날짜 형식입니다: {value}")
    return date.fromisoformat(value)


def is_valid_date(value: str) -> bool:
    try:
        parse_date(value)
        return True
    except ValueError:
        return False


def add_days(value: str, days: int) -> str:
    return (parse_date(value) + timedelta(days=days)).isoformat()


def diff_days(from_date: str, to_date: str) -> int:
    return (parse_date(to_date) - parse_date(from_date)).days


def start_of_week(value: str) -> str:
    """그 주의 월요일. (PRD 가정: 주는 월요일 시작)"""
    d = parse_date(value)
    return (d - timedelta(days=d.weekday())).isoformat()


def end_of_week(value: str) -> str:
    return add_days(start_of_week(value), 6)


WEEKDAY_KO = ["월", "화", "수", "목", "금", "토", "일"]


def format_korean_date(value: str) -> str:
    d = parse_date(value)
    return f"{d.year}년 {d.month}월 {d.day}일 ({WEEKDAY_KO[d.weekday()]})"


def format_short_date(value: str) -> str:
    d = parse_date(value)
    return f"{d.month}/{d.day}({WEEKDAY_KO[d.weekday()]})"


def format_dday(target: str, today: str | None = None) -> str:
    days = diff_days(today or today_str(), target)
    if days == 0:
        return "D-DAY"
    return f"D-{days}" if days > 0 else f"D+{-days}"


def now_timestamp() -> str:
    """생성/수정 '시각'. 달력 날짜가 아니므로 여기서만 datetime을 쓴다."""
    return datetime.now().isoformat(timespec="seconds")


def fmt_num(value: float) -> str:
    """120.0 → '120', 15.5 → '15.5'"""
    if value == int(value):
        return str(int(value))
    return f"{value:g}"


# ===========================================================================
# 데이터 모델
# ===========================================================================

AlertStatus = Literal["safe", "warning", "overdue"]
ScheduleStatus = Literal["scheduled", "completed", "cancelled"]
StockChangeType = Literal["incoming", "adjustment", "consumption"]


@dataclass
class Material:
    id: str
    name: str
    unit: str
    current_stock: float
    daily_usage: float
    lead_time_days: int
    safety_stock: float
    alert_lead_days: int
    supplier: str = ""
    memo: str = ""
    created_at: str = ""
    updated_at: str = ""


@dataclass
class IncomingSchedule:
    id: str
    material_id: str
    expected_date: str
    quantity: float
    status: ScheduleStatus = "scheduled"
    completed_at: str | None = None
    memo: str = ""
    created_at: str = ""
    updated_at: str = ""


@dataclass
class StockChangeLog:
    id: str
    material_id: str
    date: str
    type: StockChangeType
    delta: float
    stock_after: float
    reason: str = ""
    created_at: str = ""


@dataclass
class ThresholdResult:
    date: str | None
    days_from_today: int | None
    is_already_below: bool


@dataclass
class MaterialStatus:
    material_id: str
    depletion_date: str | None
    days_until_depletion: int | None
    reorder_date: str | None
    days_until_reorder: int | None
    reorder_threshold: float
    is_overdue: bool
    alert_status: AlertStatus


def new_id() -> str:
    return str(uuid4())


# ===========================================================================
# 저장소 — JSON 파일 (localStorage 대체)
# ===========================================================================
#
# 브라우저 localStorage 대신 프로세스 옆의 JSON 파일 하나를 쓴다.
# 규모(원료 수십 종)에서는 "전체 읽기 → 수정 → 전체 쓰기"가 인덱싱보다 낫다.
#
# 알려진 제약(PRD §7.4와 동일): 여러 사용자가 동시에 접속하면 마지막 저장이 이긴다.
# v1은 단일 사용자 전제로 수용한다.


class StorageError(RuntimeError):
    """저장 실패. 호출부가 사용자에게 알리고 이전 상태를 유지해야 한다."""


def _empty_store() -> dict[str, list]:
    return {"materials": [], "schedules": [], "logs": []}


def load_store() -> dict[str, list]:
    """
    파일이 깨져 있어도 앱이 죽으면 안 된다 (NF-04).
    파싱 실패 시 빈 상태로 복구하고 원본은 .corrupt로 옮겨 되살릴 여지를 남긴다.
    """
    if not STORE_PATH.exists():
        return _empty_store()
    try:
        raw = json.loads(STORE_PATH.read_text(encoding="utf-8"))
        if not isinstance(raw, dict):
            raise ValueError("최상위가 객체가 아닙니다")
        return {
            "materials": list(raw.get("materials", [])),
            "schedules": list(raw.get("schedules", [])),
            "logs": list(raw.get("logs", [])),
        }
    except Exception:  # noqa: BLE001
        logger.exception("저장 파일을 읽지 못했습니다. 빈 상태로 시작합니다.")
        try:
            STORE_PATH.rename(STORE_PATH.with_suffix(".json.corrupt"))
        except OSError:
            pass
        return _empty_store()


def save_store(store: dict[str, list]) -> None:
    try:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        # 임시 파일에 쓰고 교체 — 저장 중 죽어도 기존 파일이 반쯤 망가지지 않는다.
        tmp = STORE_PATH.with_suffix(".json.tmp")
        tmp.write_text(json.dumps(store, ensure_ascii=False, indent=2), encoding="utf-8")
        tmp.replace(STORE_PATH)
    except OSError as error:
        raise StorageError(f"저장에 실패했습니다: {error}") from error


def get_store() -> dict[str, list]:
    """세션 캐시. 매 rerun마다 파일을 읽지 않는다."""
    if "store" not in st.session_state:
        st.session_state.store = load_store()
    return st.session_state.store


def commit(store: dict[str, list]) -> None:
    """변경을 파일에 반영하고 RAG 인덱스를 낡음 표시한다."""
    save_store(store)
    st.session_state.store = store
    st.session_state.data_version = st.session_state.get("data_version", 0) + 1


def materials() -> list[Material]:
    return [Material(**m) for m in get_store()["materials"]]


def schedules() -> list[IncomingSchedule]:
    return [IncomingSchedule(**s) for s in get_store()["schedules"]]


def logs() -> list[StockChangeLog]:
    return [StockChangeLog(**l) for l in get_store()["logs"]]


def find_material(material_id: str) -> Material | None:
    return next((m for m in materials() if m.id == material_id), None)


def schedules_for(material_id: str, all_schedules: list[IncomingSchedule] | None = None):
    source = all_schedules if all_schedules is not None else schedules()
    return [s for s in source if s.material_id == material_id]


# ===========================================================================
# 도메인 계산 — 전부 순수 함수
# ===========================================================================
#
# 화면은 여기 있는 계산을 절대 재구현하지 않는다.
# D-Day·발주일·알림 등급이 필요하면 evaluate_material()을 부른다.

SIMULATION_LIMIT_DAYS = 365
DEFAULT_PROJECTION_DAYS = 60

# 차트 창을 자동으로 정할 때의 하한/상한.
# 하한이 없으면 오늘 소진되는 원료의 그래프가 점 두 개짜리가 되고,
# 상한이 없으면 재고가 아주 많은 원료에서 365일짜리 평평한 선이 나온다.
MIN_PROJECTION_DAYS = 14
MAX_PROJECTION_DAYS = 180


def build_incoming_by_date(items: list[IncomingSchedule]) -> dict[str, float]:
    """
    날짜 → 그 날 입고될 총 수량.

    INV-1: status == 'scheduled'인 일정만 집계한다.
    'completed'가 되는 순간 그 수량은 이미 Material.current_stock에 합산됐으므로
    (complete_schedule 참조) 예정일에 또 더하면 이중 계상이 된다.
    실제로 "예정일보다 일찍 완료 처리한 입고가 두 번 계산되는" 버그가 있었다.
    이 필터를 절대 넓히지 말 것.
    """
    by_date: dict[str, float] = {}
    for item in items:
        if item.status != "scheduled":
            continue
        if not np.isfinite(item.quantity) or item.quantity <= 0:
            continue
        by_date[item.expected_date] = by_date.get(item.expected_date, 0.0) + item.quantity
    return by_date


def simulate_to_threshold(
    material: Material,
    incoming_by_date: dict[str, float],
    threshold: float,
    today: str,
) -> ThresholdResult:
    """
    오늘부터 하루씩 전진하며 재고가 threshold 이하가 되는 첫 날을 찾는다.

    하루의 순서: 입고 가산 → 사용량 차감 → 임계값 판정.
    소진일에 딱 맞춰 도착하는 입고(JIT)를 결품으로 오판하지 않기 위해서다.

    나눗셈 한 번(재고 ÷ 일사용량)으로 대신할 수 없는 이유: 입고가 특정 날짜에만
    꽂히므로 재고 곡선이 톱니 모양이 된다.
    """
    if material.current_stock <= threshold:
        return ThresholdResult(date=today, days_from_today=0, is_already_below=True)

    if not np.isfinite(material.daily_usage) or material.daily_usage <= 0:
        # 사용량이 0 이하면 재고가 줄지 않는다. (검증이 막지만 방어적으로)
        return ThresholdResult(date=None, days_from_today=None, is_already_below=False)

    stock = material.current_stock
    for day in range(1, SIMULATION_LIMIT_DAYS + 1):  # INV-3
        current = add_days(today, day)
        stock += incoming_by_date.get(current, 0.0)
        stock -= material.daily_usage
        if stock <= threshold:
            return ThresholdResult(date=current, days_from_today=day, is_already_below=False)

    return ThresholdResult(date=None, days_from_today=None, is_already_below=False)


def simulate_depletion(
    material: Material, items: list[IncomingSchedule], today: str | None = None
) -> ThresholdResult:
    """재고 소진 예정일(D-Day). 임계값 0으로 시뮬레이션한다."""
    return simulate_to_threshold(
        material, build_incoming_by_date(items), 0.0, today or today_str()
    )


def reorder_threshold(material: Material) -> float:
    """리드타임 동안 쓸 양까지 미리 확보해 둬야 안전재고를 깨지 않는다."""
    return material.safety_stock + material.lead_time_days * material.daily_usage


def calculate_reorder_date(
    material: Material, items: list[IncomingSchedule], today: str | None = None
) -> tuple[ThresholdResult, float]:
    """발주 필요일 — 지금 발주해야 리드타임 안에 안전재고를 지킬 수 있는 마지막 날."""
    threshold = reorder_threshold(material)
    result = simulate_to_threshold(
        material, build_incoming_by_date(items), threshold, today or today_str()
    )
    return result, threshold


def evaluate_material(
    material: Material, items: list[IncomingSchedule], today: str | None = None
) -> MaterialStatus:
    """
    한 원료의 상태를 한 번에 계산한다.
    대시보드 / 목록 / 상세 / 리포트 / RAG 청크가 전부 이 함수를 쓴다.
    같은 수치를 두 군데서 따로 계산하면 반드시 어긋난다.

    알림 경계 규칙 (PRD §8.2):
      - 오늘 재고가 이미 임계값 이하(=)면 overdue. 오늘이 마지막 발주일이므로 즉시 조치 대상.
      - 남은 일수가 alert_lead_days와 정확히 같으면 warning (경계 포함).
      - 365일 내 발주가 필요 없으면 safe.
    """
    today = today or today_str()
    depletion = simulate_depletion(material, items, today)
    reorder, threshold = calculate_reorder_date(material, items, today)

    if reorder.is_already_below:
        alert: AlertStatus = "overdue"
    elif (
        reorder.days_from_today is not None
        and reorder.days_from_today <= material.alert_lead_days
    ):
        alert = "warning"
    else:
        alert = "safe"

    return MaterialStatus(
        material_id=material.id,
        depletion_date=depletion.date,
        days_until_depletion=depletion.days_from_today,
        reorder_date=reorder.date,
        days_until_reorder=reorder.days_from_today,
        reorder_threshold=threshold,
        is_overdue=reorder.is_already_below,
        alert_status=alert,
    )


def project_stock(
    material: Material,
    items: list[IncomingSchedule],
    today: str | None = None,
    days: int = DEFAULT_PROJECTION_DAYS,
) -> list[dict[str, Any]]:
    """
    차트용 일별 재고 투영.
    simulate_to_threshold와 달리 임계값에서 멈추지 않고 days일 전체를 돌며,
    표시용이므로 0에서 바닥을 친다(음수 재고는 현실에 없다).
    """
    today = today or today_str()
    incoming_by_date = build_incoming_by_date(items)
    usage = material.daily_usage if material.daily_usage > 0 else 0.0

    points = [{"date": today, "stock": max(0.0, material.current_stock), "incoming": 0.0}]
    stock = material.current_stock

    for day in range(1, days + 1):
        current = add_days(today, day)
        incoming = incoming_by_date.get(current, 0.0)
        stock = stock + incoming - usage
        stock = max(0.0, stock)
        points.append({"date": current, "stock": stock, "incoming": incoming})

    return points


def choose_projection_days(
    material: Material,
    items: list[IncomingSchedule],
    status: MaterialStatus,
    today: str,
) -> int:
    """
    차트에 보여줄 기간을 데이터에 맞춰 고른다.

    60일 고정으로 두면 8일 만에 소진되는 원료는 그래프의 85%가 바닥에 붙은
    평평한 선이 되어 정작 봐야 할 하강 구간이 뭉개진다.

    창에 반드시 들어가야 하는 것:
      - 소진 예정일 (가장 중요한 지점)
      - 발주 필요일
      - 앞으로 들어올 입고 예정일 — 소진 **이후**에 도착하는 입고도 포함해야
        "이 납기는 이미 늦다"가 눈에 보인다.
    거기에 여백을 조금 더해 마커가 오른쪽 끝에 붙지 않게 한다.
    """
    milestones: list[int] = []

    if status.days_until_depletion is not None:
        milestones.append(status.days_until_depletion)
    if status.days_until_reorder is not None:
        milestones.append(status.days_until_reorder)

    for item in items:
        if item.status != "scheduled":
            continue
        offset = diff_days(today, item.expected_date)
        if 0 < offset <= MAX_PROJECTION_DAYS:
            milestones.append(offset)

    if not milestones:
        # 소진도 발주도 1년 내에 없고 입고 예정도 없는 원료. 기본 창으로 둔다.
        return DEFAULT_PROJECTION_DAYS

    span = max(milestones)
    # 여백은 span의 1/4, 최소 7일. 정수 나눗셈을 쓰는 이유는 React 판(src/)과 값을
    # 정확히 맞추기 위해서다 — JS Math.round(2.5)=3, Python round(2.5)=2 라서
    # 반올림을 쓰면 두 구현의 창 길이가 하루씩 어긋난다.
    padded = span + max(7, span // 4)
    return max(MIN_PROJECTION_DAYS, min(MAX_PROJECTION_DAYS, padded))


def incoming_quantity_in_range(
    items: list[IncomingSchedule], from_date: str, to_date: str
) -> float:
    """
    from~to(포함) 사이의 입고 수량 합. 주간 리포트용.
    취소분은 뺀다. 완료분은 "그 주에 실제로 들어온 양"이므로 포함한다.
    """
    total = 0.0
    for item in items:
        if item.status == "cancelled":
            continue
        if diff_days(from_date, item.expected_date) >= 0 and diff_days(item.expected_date, to_date) >= 0:
            total += item.quantity
    return total


def is_depleting_within(status: MaterialStatus, within_days: int) -> bool:
    return status.days_until_depletion is not None and status.days_until_depletion <= within_days


def sort_key_by_alert(status: MaterialStatus) -> tuple[int, int]:
    """overdue → warning → safe, 같은 등급이면 급한 순."""
    return (
        STATUS_ORDER[status.alert_status],
        status.days_until_reorder if status.days_until_reorder is not None else 10**9,
    )


# ===========================================================================
# 검증
# ===========================================================================
#
# INV-4 적응 노트:
# 자바스크립트 판에서는 폼의 숫자 필드를 문자열로 들고 다녔다. 키 입력마다 Number()로
# 바꾸면 ''가 조용히 0이 되어 "필수" 검증을 통과해버리기 때문이다.
# Streamlit의 st.number_input은 애초에 빈 값이 될 수 없어서 그 함정은 없지만,
# **기본값 0이 그대로 진짜 데이터로 저장되는** 같은 성질의 함정이 남는다.
# 그래서 daily_usage > 0을 여기서 명시적으로 막는다. 이게 INV-4의 실질이다.


def validate_material(
    name: str,
    unit: str,
    current_stock: float,
    daily_usage: float,
    lead_time_days: int,
    safety_stock: float,
    alert_lead_days: int,
    memo: str,
    existing_names: list[str],
) -> list[str]:
    """에러 메시지 목록. 비어 있으면 통과."""
    errors: list[str] = []

    clean_name = name.strip()
    if not clean_name:
        errors.append("원료명을 입력하세요.")
    elif len(clean_name) > 50:
        errors.append("원료명은 50자 이하여야 합니다.")
    elif any(n.strip().lower() == clean_name.lower() for n in existing_names):
        errors.append("같은 이름의 원료가 이미 있습니다.")

    if not unit.strip():
        errors.append("단위를 입력하세요. (예: kg, L, EA)")

    if current_stock < 0:
        errors.append("현재 재고는 0 이상이어야 합니다.")

    if daily_usage <= 0:
        # 0이면 재고가 줄지 않아 소진일 계산 자체가 성립하지 않는다 (PRD §13).
        errors.append("일일 평균 사용량은 0보다 커야 합니다.")

    if lead_time_days < 0:
        errors.append("리드타임은 0 이상이어야 합니다.")

    if safety_stock < 0:
        errors.append("안전재고는 0 이상이어야 합니다.")

    if alert_lead_days < 0:
        errors.append("알림 선행일은 0 이상이어야 합니다.")

    if len(memo) > 500:
        errors.append("메모는 500자 이하여야 합니다.")

    return errors


def validate_schedule(expected_date: str, quantity: float, memo: str) -> list[str]:
    errors: list[str] = []
    if not is_valid_date(expected_date):
        errors.append("입고 예정일이 올바르지 않습니다.")
    if quantity <= 0:
        errors.append("입고 수량은 0보다 커야 합니다.")
    if len(memo) > 200:
        errors.append("비고는 200자 이하여야 합니다.")
    return errors


# ===========================================================================
# 액션 — 데이터 변경의 유일한 통로
# ===========================================================================
#
# 화면은 여기 있는 함수만 호출하고 store를 직접 건드리지 않는다.
# 여러 엔티티가 함께 바뀌는 규칙(입고 완료, 캐스케이드 삭제)을 한곳에 모아둔다.


def create_material(**kwargs) -> Material:
    timestamp = now_timestamp()
    material = Material(
        id=new_id(),
        created_at=timestamp,
        updated_at=timestamp,
        **kwargs,
    )
    store = get_store()
    store["materials"].append(asdict(material))
    commit(store)
    return material


def update_material(material_id: str, **changes) -> None:
    store = get_store()
    for item in store["materials"]:
        if item["id"] == material_id:
            item.update(changes)
            item["updated_at"] = now_timestamp()
            break
    commit(store)


def delete_material(material_id: str) -> None:
    """
    원료를 지우면 연결된 입고 예정과 변동 이력도 함께 지운다.
    남겨두면 고아 레코드가 되어 리포트·RAG 청크에 유령 데이터로 새어 나온다.
    """
    store = get_store()
    store["materials"] = [m for m in store["materials"] if m["id"] != material_id]
    store["schedules"] = [s for s in store["schedules"] if s["material_id"] != material_id]
    store["logs"] = [l for l in store["logs"] if l["material_id"] != material_id]
    commit(store)


def create_schedule(material_id: str, expected_date: str, quantity: float, memo: str = "") -> None:
    timestamp = now_timestamp()
    schedule = IncomingSchedule(
        id=new_id(),
        material_id=material_id,
        expected_date=expected_date,
        quantity=quantity,
        status="scheduled",
        memo=memo.strip(),
        created_at=timestamp,
        updated_at=timestamp,
    )
    store = get_store()
    store["schedules"].append(asdict(schedule))
    commit(store)


def update_schedule(schedule_id: str, expected_date: str, quantity: float, memo: str = "") -> None:
    store = get_store()
    for item in store["schedules"]:
        if item["id"] == schedule_id:
            item["expected_date"] = expected_date
            item["quantity"] = quantity
            item["memo"] = memo.strip()
            item["updated_at"] = now_timestamp()
            break
    commit(store)


def cancel_schedule(schedule_id: str) -> None:
    store = get_store()
    for item in store["schedules"]:
        if item["id"] == schedule_id:
            item["status"] = "cancelled"
            item["updated_at"] = now_timestamp()
            break
    commit(store)


def delete_schedule(schedule_id: str) -> None:
    store = get_store()
    store["schedules"] = [s for s in store["schedules"] if s["id"] != schedule_id]
    commit(store)


def complete_schedule(schedule_id: str, today: str | None = None) -> bool:
    """
    입고 완료 처리 — INV-1이 존재하는 이유.

    세 가지가 한 번에 일어난다:
      1. Material.current_stock += quantity   (재고에 실제 반영)
      2. schedule.status = 'completed'        (시뮬레이션 집계에서 빠짐)
      3. StockChangeLog(type='incoming') 추가 (이력)

    2번이 빠지면 재고에 더해진 수량이 예정일에 또 더해져 이중 계상된다.
    이미 completed인 일정은 무시한다 — 버튼을 두 번 눌러도 재고가 두 번 늘지 않는다.
    """
    today = today or today_str()
    store = get_store()

    schedule = next((s for s in store["schedules"] if s["id"] == schedule_id), None)
    if schedule is None or schedule["status"] != "scheduled":
        return False  # 멱등

    material = next((m for m in store["materials"] if m["id"] == schedule["material_id"]), None)
    if material is None:
        return False

    timestamp = now_timestamp()
    stock_after = material["current_stock"] + schedule["quantity"]

    material["current_stock"] = stock_after
    material["updated_at"] = timestamp

    schedule["status"] = "completed"
    schedule["completed_at"] = timestamp
    schedule["updated_at"] = timestamp

    store["logs"].append(
        asdict(
            StockChangeLog(
                id=new_id(),
                material_id=material["id"],
                date=today,
                type="incoming",
                delta=schedule["quantity"],
                stock_after=stock_after,
                reason=f"입고 완료 (예정일 {schedule['expected_date']})",
                created_at=timestamp,
            )
        )
    )
    commit(store)
    return True


def adjust_stock(material_id: str, new_stock: float, reason: str, today: str | None = None) -> bool:
    """실사/폐기 등으로 재고를 직접 바꾸고 이력을 남긴다."""
    today = today or today_str()
    store = get_store()

    material = next((m for m in store["materials"] if m["id"] == material_id), None)
    if material is None:
        return False

    delta = new_stock - material["current_stock"]
    if delta == 0:
        return False

    timestamp = now_timestamp()
    material["current_stock"] = new_stock
    material["updated_at"] = timestamp

    store["logs"].append(
        asdict(
            StockChangeLog(
                id=new_id(),
                material_id=material_id,
                date=today,
                type="adjustment",
                delta=delta,
                stock_after=new_stock,
                reason=reason.strip(),
                created_at=timestamp,
            )
        )
    )
    commit(store)
    return True


def seed_sample_data(today: str | None = None) -> None:
    """빈 화면에서 기능을 확인할 수 있게 하는 샘플 데이터."""
    today = today or today_str()
    timestamp = now_timestamp()
    store = get_store()

    specs = [
        ("글리세린", "kg", 120, 15, 7, 50, "A상사"),
        ("정제수", "L", 800, 120, 3, 300, "B케미칼"),
        ("스테아르산", "kg", 640, 8, 14, 100, "C머티리얼"),
        ("부틸렌글라이콜", "kg", 45, 6, 10, 40, "A상사"),
        ("세틸알코올", "kg", 300, 5, 5, 60, "D인터내셔널"),
    ]

    created: list[Material] = []
    for name, unit, stock, usage, lead, safety, supplier in specs:
        material = Material(
            id=new_id(),
            name=name,
            unit=unit,
            current_stock=float(stock),
            daily_usage=float(usage),
            lead_time_days=lead,
            safety_stock=float(safety),
            alert_lead_days=3,
            supplier=supplier,
            created_at=timestamp,
            updated_at=timestamp,
        )
        created.append(material)
        store["materials"].append(asdict(material))

    for material, offset, quantity, memo in [
        (created[0], 4, 200.0, "정기 발주분"),
        (created[1], 2, 1000.0, ""),
    ]:
        store["schedules"].append(
            asdict(
                IncomingSchedule(
                    id=new_id(),
                    material_id=material.id,
                    expected_date=add_days(today, offset),
                    quantity=quantity,
                    status="scheduled",
                    memo=memo,
                    created_at=timestamp,
                    updated_at=timestamp,
                )
            )
        )

    commit(store)


# ===========================================================================
# 리포트 행 생성 — 화면 표 / TSV 복사 / 엑셀이 공유하는 단 하나의 함수
# ===========================================================================

REPORT_HEADERS = [
    "원료명",
    "단위",
    "현재재고",
    "일사용량",
    "안전재고",
    "리드타임(일)",
    "소진예정일",
    "D-Day",
    "발주필요일",
    "상태",
    "주간입고예정",
]

NUMERIC_HEADERS = {"현재재고", "일사용량", "안전재고", "리드타임(일)", "주간입고예정"}


def sanitize_cell(value: str) -> str:
    """
    TSV/엑셀 셀 안전화.
    값에 탭이나 개행이 들어가면 붙여넣은 표가 통째로 깨진다.
    셀 경계를 지키는 게 원문 보존보다 중요하므로 공백으로 치환한다.
    """
    return re.sub(r"[\t\r\n]+", " ", value).strip()


@dataclass
class ReportRow:
    material: Material
    status: MaterialStatus
    weekly_incoming: float
    cells: list[str]


def build_report_rows(
    material_list: list[Material],
    schedule_list: list[IncomingSchedule],
    week_start: str,
    week_end: str,
    today: str,
    risky_only: bool = False,
) -> list[ReportRow]:
    rows: list[ReportRow] = []

    for material in material_list:
        own = schedules_for(material.id, schedule_list)
        status = evaluate_material(material, own, today)
        weekly = incoming_quantity_in_range(own, week_start, week_end)

        cells = [
            sanitize_cell(material.name),
            sanitize_cell(material.unit),
            fmt_num(material.current_stock),
            fmt_num(material.daily_usage),
            fmt_num(material.safety_stock),
            str(material.lead_time_days),
            status.depletion_date or "1년 내 없음",
            format_dday(status.depletion_date, today) if status.depletion_date else "-",
            status.reorder_date or "불필요",
            STATUS_LABELS[status.alert_status],
            fmt_num(weekly),
        ]
        rows.append(ReportRow(material, status, weekly, cells))

    if risky_only:
        rows = [r for r in rows if r.status.alert_status != "safe"]

    # 급한 것부터: overdue → warning → safe, 같은 등급이면 소진 임박순
    rows.sort(
        key=lambda r: (
            STATUS_ORDER[r.status.alert_status],
            r.status.days_until_depletion if r.status.days_until_depletion is not None else 10**9,
        )
    )
    return rows


def rows_to_tsv(rows: list[ReportRow]) -> str:
    """클립보드용. 셀에서 탭·개행은 이미 sanitize_cell이 제거했다."""
    lines = ["\t".join(REPORT_HEADERS)] + ["\t".join(r.cells) for r in rows]
    return "\n".join(lines)


def rows_to_csv(rows: list[ReportRow]) -> str:
    """
    CSV는 탭과 달리 쉼표가 셀 안에 들어올 수 있다("A상사, B상사" 같은 값).
    따옴표로 감싸고 내부 따옴표는 두 번 써서 이스케이프한다.
    """

    def escape(value: str) -> str:
        if '"' in value or "," in value:
            return '"' + value.replace('"', '""') + '"'
        return value

    lines = [",".join(escape(h) for h in REPORT_HEADERS)]
    lines += [",".join(escape(cell) for cell in row.cells) for row in rows]
    return "\n".join(lines)


def rows_to_excel(rows: list[ReportRow], week_start: str, week_end: str) -> bytes:
    """
    openpyxl로 .xlsx 생성. 헤더 굵게, 열 너비, 상태별 행 색상.

    (자바스크립트 판에서 SheetJS를 금지했던 이유는 npm 배포 빌드의 보안 권고 때문이고,
     파이썬 openpyxl과는 무관하다.)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    fills = {
        "발주 지연": PatternFill("solid", fgColor="FDE7E7"),
        "발주 임박": PatternFill("solid", fgColor="FFF4E0"),
        "안전": PatternFill("solid", fgColor="EFF7EE"),
    }

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "주간 수급"

    # 제목 줄
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(REPORT_HEADERS))
    title = sheet.cell(row=1, column=1, value=f"주간 수급 리포트  {week_start} ~ {week_end}")
    title.font = Font(bold=True, size=13)
    title.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 24

    # 헤더
    header_fill = PatternFill("solid", fgColor="EFF1F5")
    thin_bottom = Border(bottom=Side(style="thin", color="CBD2DA"))
    for column, header in enumerate(REPORT_HEADERS, start=1):
        cell = sheet.cell(row=2, column=column, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_bottom

    # 본문
    numeric_columns = {3, 4, 5, 6, 11}  # 1-based
    for row_index, row in enumerate(rows, start=3):
        status_text = row.cells[9]
        fill = fills.get(status_text)
        for column, value in enumerate(row.cells, start=1):
            if column in numeric_columns:
                try:
                    value = float(value)
                    if value == int(value):
                        value = int(value)
                except ValueError:
                    pass
            cell = sheet.cell(row=row_index, column=column, value=value)
            if fill:
                cell.fill = fill
            if column == 10 and status_text != "안전":
                cell.font = Font(bold=True)

    widths = [18, 8, 12, 12, 12, 12, 14, 10, 14, 12, 14]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.freeze_panes = "A3"
    sheet.auto_filter.ref = f"A2:{get_column_letter(len(REPORT_HEADERS))}2"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# ===========================================================================
# RAG — 청킹 / 임베딩 / FAISS / MMR
# ===========================================================================
#
# 토큰 효율이 이 절의 존재 이유다. 전체 데이터를 프롬프트에 통째로 넣으면
# 원료 40종 기준 8,000~12,000 토큰이 든다. 아래 6단계를 거치면 평균 1,500 토큰 이하가 된다.
#
#   1. 문장 단위 청킹              → build_chunks
#   2. 정규화 + IndexFlatIP        → 코사인 유사도
#   3. top-20 → MMR(λ=0.5) → 5    → 중복 제거
#   4. 유사도 0.30 미만 폐기       → 관련 없는 청크 배제
#   5. 컨텍스트 1,200 토큰 상한    → 하드 리밋
#   6. 해시 기반 증분 임베딩 + 캐시 → 비용


@dataclass
class Chunk:
    key: str
    type: str
    text: str
    label: str
    material_id: str | None = None
    material_name: str | None = None
    date: str | None = None

    @property
    def content_hash(self) -> str:
        """텍스트가 바뀌었는지 판정하는 지문. 같으면 기존 임베딩을 재사용한다."""
        return hashlib.sha256(self.text.encode("utf-8")).hexdigest()


def build_chunks(
    material_list: list[Material],
    schedule_list: list[IncomingSchedule],
    log_list: list[StockChangeLog],
    today: str,
) -> list[Chunk]:
    """
    스냅샷 → 검색 가능한 청크.

    원료 1건을 통째로 한 덩어리로 만들면 "입고 언제 와?"라는 질문에도
    재고·안전재고·리드타임이 전부 딸려 온다. 그래서 원료 요약 / 입고 예정 /
    변동 이력을 각각 다른 청크로 쪼갠다.

    AI-02: D-Day와 발주일은 evaluate_material()이 계산한 값을 문장에 그대로 넣는다.
    LLM은 이 수치를 계산하지 않고 인용만 한다.
    """
    chunks: list[Chunk] = []
    name_by_id = {m.id: m.name for m in material_list}
    unit_by_id = {m.id: m.unit for m in material_list}

    statuses = {
        m.id: evaluate_material(m, schedules_for(m.id, schedule_list), today)
        for m in material_list
    }

    # 전체 현황 요약 — "몇 종이야?", "위험한 게 몇 개야?" 류 질문은
    # 개별 원료 청크로는 답이 안 나온다(5개만 검색되니 전체를 셀 수 없다).
    if material_list:
        overdue = [m.name for m in material_list if statuses[m.id].alert_status == "overdue"]
        warning = [m.name for m in material_list if statuses[m.id].alert_status == "warning"]
        safe_count = len(material_list) - len(overdue) - len(warning)
        summary = (
            f"기준일 {today}. 전체 원료 {len(material_list)}종. "
            f"발주 지연 {len(overdue)}종" + (f" ({', '.join(overdue)}). " if overdue else ". ")
            + f"발주 임박 {len(warning)}종" + (f" ({', '.join(warning)}). " if warning else ". ")
            + f"안전 {safe_count}종."
        )
        chunks.append(Chunk(key="overview:all", type="policy", text=summary, label="전체 현황 요약"))

    for material in material_list:
        status = statuses[material.id]
        parts = [
            f"{material.name}({material.unit}): "
            f"현재재고 {fmt_num(material.current_stock)}{material.unit}, "
            f"일사용량 {fmt_num(material.daily_usage)}{material.unit}, "
            f"안전재고 {fmt_num(material.safety_stock)}{material.unit}, "
            f"리드타임 {material.lead_time_days}일."
        ]
        if material.supplier:
            parts.append(f"공급업체 {material.supplier}.")

        if status.depletion_date:
            days = status.days_until_depletion
            marker = f"(D-{days})" if days else "(오늘)"
            parts.append(f"소진 예정일 {status.depletion_date} {marker}.")
        else:
            parts.append("1년 내 소진 예정 없음.")

        if status.reorder_date:
            parts.append(f"발주 필요일 {status.reorder_date}.")
        else:
            parts.append("1년 내 발주 불필요.")

        parts.append(f"상태: {STATUS_LABELS[status.alert_status]}.")
        if material.memo:
            parts.append(f"메모: {material.memo}")

        chunks.append(
            Chunk(
                key=f"material:{material.id}",
                type="material",
                text=" ".join(parts),
                label=f"{material.name} / 원료요약",
                material_id=material.id,
                material_name=material.name,
                date=status.reorder_date,
            )
        )

    for schedule in schedule_list:
        # 고아 레코드(원료가 지워진 일정)와 취소분은 인덱싱하지 않는다.
        # 답이 될 일이 없는데 토큰만 먹는다.
        if schedule.material_id not in name_by_id or schedule.status == "cancelled":
            continue
        name = name_by_id[schedule.material_id]
        unit = unit_by_id[schedule.material_id]
        text = (
            f"{name} 입고 일정: {schedule.expected_date}, "
            f"수량 {fmt_num(schedule.quantity)}{unit}, "
            f"상태 {SCHEDULE_STATUS_LABELS[schedule.status]}."
        )
        if schedule.memo:
            text += f" 비고: {schedule.memo}"
        chunks.append(
            Chunk(
                key=f"schedule:{schedule.id}",
                type="schedule",
                text=text,
                label=f"{name} / 입고예정 {schedule.expected_date}",
                material_id=schedule.material_id,
                material_name=name,
                date=schedule.expected_date,
            )
        )

    # 이력은 최근 200건만. 오래된 로그는 질문에 거의 안 쓰이는데 인덱스만 키운다.
    recent_logs = sorted(log_list, key=lambda l: l.created_at, reverse=True)[:200]
    for log in recent_logs:
        if log.material_id not in name_by_id:
            continue
        name = name_by_id[log.material_id]
        unit = unit_by_id[log.material_id]
        sign = "+" if log.delta >= 0 else ""
        text = (
            f"{name} {log.date} {LOG_TYPE_LABELS[log.type]} {sign}{fmt_num(log.delta)}{unit}, "
            f"변동 후 {fmt_num(log.stock_after)}{unit}."
        )
        if log.reason:
            text += f" 사유: {log.reason}"
        chunks.append(
            Chunk(
                key=f"log:{log.id}",
                type="log",
                text=text,
                label=f"{name} / 변동이력 {log.date}",
                material_id=log.material_id,
                material_name=name,
                date=log.date,
            )
        )

    return chunks


def estimate_tokens(text: str) -> int:
    """
    토큰 수 근사. tiktoken이 있으면 쓰고, 없으면 휴리스틱으로 넘어간다.
    한국어는 대략 1토큰 ≈ 1.3~1.8자라 보수적으로 '문자수 / 1.3'으로 잡는다.
    상한을 지키는 게 목적이므로 과소평가보다 과대평가가 안전하다.
    """
    try:
        import tiktoken

        return len(tiktoken.get_encoding("cl100k_base").encode(text))
    except Exception:  # noqa: BLE001 - tiktoken은 선택적 의존성
        return int(len(text) / 1.3) + 1


def normalize(vectors: np.ndarray) -> np.ndarray:
    """L2 정규화. 정규화된 벡터의 내적 = 코사인 유사도라 IndexFlatIP를 쓸 수 있다."""
    norms = np.linalg.norm(vectors, axis=1, keepdims=True)
    norms[norms == 0] = 1.0
    return vectors / norms


# --- 임베딩 캐시 (해시 → 벡터). 디스크에 남아 재시작해도 재사용된다. -------------

_cache_lock = threading.Lock()


def load_embed_cache() -> dict[str, np.ndarray]:
    if not EMBED_CACHE_PATH.exists():
        return {}
    try:
        with np.load(EMBED_CACHE_PATH, allow_pickle=False) as archive:
            hashes = archive["hashes"]
            vectors = archive["vectors"]
        if len(hashes) != vectors.shape[0]:
            return {}
        return {str(h): vectors[i] for i, h in enumerate(hashes)}
    except Exception:  # noqa: BLE001 - 손상된 캐시 때문에 앱이 죽으면 안 된다
        logger.exception("임베딩 캐시를 읽지 못했습니다. 새로 만듭니다.")
        return {}


def save_embed_cache(cache: dict[str, np.ndarray]) -> None:
    try:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        items = list(cache.items())[-EMBED_CACHE_MAX:]  # 무한 증식 방지
        if not items:
            return
        hashes = np.array([k for k, _ in items])
        vectors = np.stack([v for _, v in items]).astype("float32")
        np.savez_compressed(EMBED_CACHE_PATH, hashes=hashes, vectors=vectors)
    except Exception:  # noqa: BLE001
        logger.exception("임베딩 캐시 저장 실패 (동작에는 지장 없음)")


# --- OpenAI ---------------------------------------------------------------


class LlmError(RuntimeError):
    """사용자에게 그대로 보여줄 수 있는 메시지를 담는다."""


@st.cache_resource(show_spinner=False)
def get_openai_client(api_key: str):
    from openai import OpenAI

    return OpenAI(api_key=api_key)


def translate_error(error: Exception) -> LlmError:
    """OpenAI SDK 예외를 사용자용 메시지로."""
    name = type(error).__name__
    text = str(error)

    if "AuthenticationError" in name or "invalid_api_key" in text:
        return LlmError("OpenAI 인증에 실패했습니다. .env의 OPENAI_API_KEY를 확인해 주세요.")
    if "RateLimitError" in name or "rate_limit" in text:
        return LlmError("OpenAI 요청 한도를 초과했습니다. 잠시 후 다시 시도해 주세요.")
    if "NotFoundError" in name or "model_not_found" in text or "does not exist" in text:
        return LlmError(
            f"모델을 찾을 수 없습니다: {get_chat_model()} / {get_embedding_model()}. "
            ".env의 모델명을 확인해 주세요."
        )
    if "APITimeoutError" in name or "Timeout" in name:
        return LlmError("OpenAI 응답 시간이 초과되었습니다. 잠시 후 다시 시도해 주세요.")
    if "APIConnectionError" in name:
        return LlmError("OpenAI에 연결하지 못했습니다. 네트워크를 확인해 주세요.")

    logger.exception("OpenAI 호출 실패")
    return LlmError(f"OpenAI 호출에 실패했습니다: {text[:200]}")


def embed_texts(texts: list[str]) -> np.ndarray:
    """최대 100개씩 배치로 호출한다. 한 건씩 부르면 왕복이 청크 수만큼 늘어난다."""
    if not texts:
        return np.zeros((0, EMBEDDING_DIM), dtype="float32")

    api_key = get_api_key()
    if not api_key:
        raise LlmError(".env에 OPENAI_API_KEY가 설정되지 않았습니다.")

    client = get_openai_client(api_key)
    model = get_embedding_model()
    vectors: list[list[float]] = []

    try:
        for start in range(0, len(texts), RAG_EMBED_BATCH):
            batch = texts[start : start + RAG_EMBED_BATCH]
            response = client.embeddings.create(model=model, input=batch)
            # 응답 순서가 요청 순서와 같다는 보장이 문서에 있지만 index로 정렬해 확실히 한다.
            ordered = sorted(response.data, key=lambda item: item.index)
            vectors.extend(item.embedding for item in ordered)
    except LlmError:
        raise
    except Exception as error:  # noqa: BLE001
        raise translate_error(error) from error

    result = np.array(vectors, dtype="float32")
    if result.shape[1] != EMBEDDING_DIM:
        raise LlmError(
            f"임베딩 차원이 예상과 다릅니다: {result.shape[1]} (기대 {EMBEDDING_DIM}). "
            "OPENAI_EMBEDDING_MODEL을 확인하세요."
        )
    return result


# --- 인덱스 ---------------------------------------------------------------


@dataclass
class VectorIndex:
    chunks: list[Chunk] = field(default_factory=list)
    vectors: np.ndarray = field(default_factory=lambda: np.zeros((0, EMBEDDING_DIM), dtype="float32"))
    faiss_index: Any = None
    data_version: int = -1

    @property
    def size(self) -> int:
        return len(self.chunks)


def build_index(chunks: list[Chunk]) -> tuple[VectorIndex, int, int]:
    """
    청크 목록으로 FAISS 인덱스를 만든다.
    내용 해시가 캐시에 있으면 임베딩을 재사용한다 — 재고 하나 바꿨다고
    전체를 다시 임베딩하면 비용이 그대로 나간다.

    :returns: (인덱스, 새로 임베딩한 수, 재사용한 수)
    """
    import faiss

    with _cache_lock:
        cache = load_embed_cache()

    missing = [c for c in chunks if c.content_hash not in cache]
    if missing:
        embedded = embed_texts([c.text for c in missing])
        for chunk, vector in zip(missing, embedded):
            cache[chunk.content_hash] = vector
        with _cache_lock:
            save_embed_cache(cache)

    if not chunks:
        return VectorIndex(), 0, 0

    matrix = normalize(np.stack([cache[c.content_hash] for c in chunks]).astype("float32"))
    faiss_index = faiss.IndexFlatIP(matrix.shape[1])
    faiss_index.add(matrix)

    return (
        VectorIndex(chunks=list(chunks), vectors=matrix, faiss_index=faiss_index),
        len(missing),
        len(chunks) - len(missing),
    )


def mmr_select(
    candidate_vectors: np.ndarray,
    candidate_scores: list[float],
    top_k: int,
    lambda_param: float,
) -> list[int]:
    """
    Maximal Marginal Relevance — 관련성과 다양성을 동시에 본다.

    이게 없으면 "글리세린 어때?" 한 마디에 글리세린 설명 청크 5개가 뽑혀
    같은 내용을 다섯 번 프롬프트에 싣게 된다.

        score(c) = λ·sim(q, c) − (1−λ)·max(sim(c, 이미 뽑힌 것))
    """
    if not candidate_scores:
        return []

    top_k = min(top_k, len(candidate_scores))
    remaining = list(range(len(candidate_scores)))

    first = max(remaining, key=lambda i: candidate_scores[i])
    selected = [first]
    remaining.remove(first)

    while len(selected) < top_k and remaining:
        best_index, best_score = None, float("-inf")
        for index in remaining:
            # 벡터가 정규화돼 있으니 내적 = 코사인
            redundancy = max(
                float(np.dot(candidate_vectors[index], candidate_vectors[chosen]))
                for chosen in selected
            )
            score = lambda_param * candidate_scores[index] - (1 - lambda_param) * redundancy
            if score > best_score:
                best_index, best_score = index, score
        if best_index is None:
            break
        selected.append(best_index)
        remaining.remove(best_index)

    return selected


WEEK_PATTERNS = {
    "이번 주": (0, 6),
    "이번주": (0, 6),
    "금주": (0, 6),
    "다음 주": (7, 13),
    "다음주": (7, 13),
    "차주": (7, 13),
    "오늘": (0, 0),
    "내일": (1, 1),
}


def date_window(question: str, today: str) -> tuple[str, str] | None:
    """'이번 주', '다음 주' 같은 표현을 날짜 범위로."""
    try:
        base = parse_date(today)
    except ValueError:
        return None

    for pattern, (start_offset, end_offset) in WEEK_PATTERNS.items():
        if pattern in question:
            if "주" in pattern:
                monday = base - timedelta(days=base.weekday())
                return (
                    (monday + timedelta(days=start_offset)).isoformat(),
                    (monday + timedelta(days=end_offset)).isoformat(),
                )
            return (
                (base + timedelta(days=start_offset)).isoformat(),
                (base + timedelta(days=end_offset)).isoformat(),
            )
    return None


def build_prefilter(question: str, today: str, index: VectorIndex) -> set[int] | None:
    """
    검색 후보를 좁히는 인덱스 집합. 좁힐 근거가 없으면 None(전체 검색).

    좁히더라도 전체 현황 요약은 항상 남긴다 — "몇 종이야?" 같은 질문의 유일한 답이다.
    """
    names = {c.material_name for c in index.chunks if c.material_name}
    mentioned = {n for n in names if n and n in question}
    window = date_window(question, today)

    if not mentioned and not window:
        return None

    allowed: set[int] = set()
    for position, chunk in enumerate(index.chunks):
        if chunk.type == "policy":
            allowed.add(position)
            continue
        if mentioned and chunk.material_name not in mentioned:
            continue
        if window and chunk.date and chunk.type != "material":
            # 원료 요약은 날짜와 무관하게 유용하므로 창 밖이어도 남긴다.
            if not (window[0] <= chunk.date <= window[1]):
                continue
        allowed.add(position)

    # 필터가 너무 공격적이어서 아무것도 안 남으면 전체 검색으로 되돌린다.
    return allowed or None


@dataclass
class Retrieved:
    chunk: Chunk
    score: float


def assemble_context(retrieved: list[Retrieved], max_tokens: int) -> tuple[str, list[Retrieved]]:
    """
    컨텍스트 문자열과 실제로 포함된 청크를 반환한다.
    상한을 넘기면 유사도가 낮은 것부터 잘라낸다.
    """
    ordered = sorted(retrieved, key=lambda r: r.score, reverse=True)
    lines: list[str] = []
    included: list[Retrieved] = []
    used = 0

    for item in ordered:
        line = f"[{len(included) + 1}] ({item.chunk.label}) {item.chunk.text}"
        cost = estimate_tokens(line)
        # 아무것도 없이 LLM을 부르는 것보다는 첫 청크 하나는 넣는 편이 낫다.
        if used + cost > max_tokens and included:
            break
        lines.append(line)
        included.append(item)
        used += cost

    return "\n".join(lines), included


SYSTEM_PROMPT = """당신은 원료 재고 관리 어시스턴트다.
아래 <context>에 있는 정보만 근거로 답한다. 없는 내용은 추측하지 않는다.
숫자와 날짜는 context에 적힌 값을 그대로 인용한다. 직접 계산하지 않는다.
답변은 한국어로, 3문장 이내로 간결하게. 항목이 여러 개면 표를 써도 좋다.
context에 근거가 없으면 "해당 정보를 찾지 못했습니다"라고 답한다.
재고·발주·입고와 무관한 질문에는 "재고 관련 질문에만 답할 수 있습니다"라고 답한다.
오늘 날짜: {today}"""

NO_CONTEXT_ANSWER = (
    "해당 정보를 찾지 못했습니다. 등록된 원료·입고 예정 데이터에 없는 내용이거나, "
    "질문을 조금 더 구체적으로 적어 주시면 찾아볼 수 있습니다."
)


# temperature를 거부하는 모델을 기억해 둔다.
# gpt-5.6-sol은 temperature=0에 400을 돌려준다. 이 사실을 기억하지 않으면
# 질문할 때마다 실패 왕복이 한 번씩 더 붙는다(실제로 매 호출 400 → 재시도였다).
_models_without_temperature: set[str] = set()


def generate_answer(messages: list[dict[str, str]]) -> tuple[str, int, int]:
    """
    :returns: (답변, 프롬프트 토큰, 완성 토큰)

    temperature를 고정값으로 보내면 일부 신형 모델이 400을 돌려준다.
    그래서 실패하면 파라미터를 빼고 한 번 더 시도하고, 그 결과를 기억한다.
    """
    api_key = get_api_key()
    if not api_key:
        raise LlmError(".env에 OPENAI_API_KEY가 설정되지 않았습니다.")

    client = get_openai_client(api_key)
    model = get_chat_model()

    def call(with_temperature: bool):
        kwargs: dict[str, Any] = {"model": model, "messages": messages}
        if with_temperature:
            kwargs["temperature"] = 0
        return client.chat.completions.create(**kwargs)

    supports_temperature = model not in _models_without_temperature

    try:
        try:
            response = call(supports_temperature)
        except Exception as error:  # noqa: BLE001
            if supports_temperature and "temperature" in str(error).lower():
                logger.info("%s 는 temperature를 지원하지 않습니다. 기본값으로 재시도합니다.", model)
                _models_without_temperature.add(model)
                response = call(False)
            else:
                raise
    except LlmError:
        raise
    except Exception as error:  # noqa: BLE001
        raise translate_error(error) from error

    answer = (response.choices[0].message.content or "").strip()
    if not answer:
        raise LlmError("모델이 빈 응답을 반환했습니다.")

    usage = getattr(response, "usage", None)
    return (
        answer,
        int(getattr(usage, "prompt_tokens", 0) or 0),
        int(getattr(usage, "completion_tokens", 0) or 0),
    )


def answer_question(
    question: str, history: list[dict[str, str]], today: str, index: VectorIndex
) -> dict[str, Any]:
    """질문 → 사전필터 → 벡터검색 → MMR → 토큰예산 → 답변."""
    if index.size == 0 or index.faiss_index is None:
        raise LlmError("인덱스가 비어 있습니다. '인덱스 재생성'을 먼저 실행해 주세요.")

    query = normalize(embed_texts([question]))[0]

    allowed = build_prefilter(question, today, index)

    # 사전필터가 걸리면 통과분이 적을 수 있으므로 넉넉히 뽑고 나서 거른다.
    k = index.size if allowed is not None else min(index.size, RAG_CANDIDATE_K)
    scores, positions = index.faiss_index.search(query.reshape(1, -1).astype("float32"), k)

    candidates: list[tuple[int, float]] = []
    for score, position in zip(scores[0], positions[0]):
        if position < 0:
            continue
        if allowed is not None and int(position) not in allowed:
            continue
        if float(score) < RAG_MIN_SCORE:  # 유사도 임계값 미만은 버린다
            continue
        candidates.append((int(position), float(score)))
        if len(candidates) >= RAG_CANDIDATE_K:
            break

    if not candidates:
        # 근거가 없으면 지어내지 않는다. LLM을 부르지도 않는다(비용 절약).
        return {"answer": NO_CONTEXT_ANSWER, "sources": [], "prompt_tokens": 0}

    candidate_positions = [p for p, _ in candidates]
    candidate_scores = [s for _, s in candidates]
    candidate_vectors = index.vectors[candidate_positions]

    chosen = mmr_select(candidate_vectors, candidate_scores, RAG_TOP_K, RAG_MMR_LAMBDA)
    retrieved = [
        Retrieved(chunk=index.chunks[candidate_positions[i]], score=candidate_scores[i])
        for i in chosen
    ]

    context, included = assemble_context(retrieved, RAG_MAX_CONTEXT_TOKENS)
    if not included:
        return {"answer": NO_CONTEXT_ANSWER, "sources": [], "prompt_tokens": 0}

    messages: list[dict[str, str]] = [
        {"role": "system", "content": SYSTEM_PROMPT.format(today=today)}
    ]
    # 대화 이력은 직전 2턴(=4메시지)만. 전부 실으면 턴마다 프롬프트가 선형으로 커진다.
    for item in history[-4:]:
        if item.get("role") in ("user", "assistant") and item.get("content"):
            messages.append({"role": item["role"], "content": item["content"][:1000]})
    messages.append(
        {"role": "user", "content": f"<context>\n{context}\n</context>\n\n질문: {question}"}
    )

    answer, prompt_tokens, _ = generate_answer(messages)

    if prompt_tokens == 0:  # usage를 안 주는 모델 대비
        prompt_tokens = sum(estimate_tokens(m["content"]) for m in messages)

    return {
        "answer": answer,
        "sources": [
            {"id": i + 1, "label": item.chunk.label, "score": round(item.score, 4)}
            for i, item in enumerate(included)
        ],
        "prompt_tokens": prompt_tokens,
    }


# ===========================================================================
# UI 헬퍼
# ===========================================================================


def status_chip(status: AlertStatus) -> str:
    """NF-07: 색상만으로 구분하지 않는다. 아이콘 + 텍스트를 함께 쓴다."""
    return f"{STATUS_ICONS[status]} {STATUS_LABELS[status]}"


def dday_text(target: str | None, today: str, empty: str = "-") -> str:
    if not target:
        return empty
    return f"{target} ({format_dday(target, today)})"


def stock_chart(
    material: Material,
    own_schedules: list[IncomingSchedule],
    status: MaterialStatus,
    today: str,
    days: int,
):
    """재고 추이 예측 — 라인 + 안전재고 기준선 + 소진일 마커 + 입고 지점."""
    points = project_stock(material, own_schedules, today, days)
    frame = pd.DataFrame(points)
    frame["date"] = pd.to_datetime(frame["date"])

    line = (
        alt.Chart(frame)
        .mark_line(color=COLOR_LINE, strokeWidth=2)
        .encode(
            x=alt.X("date:T", title=None),
            y=alt.Y("stock:Q", title=f"재고 ({material.unit})"),
            tooltip=[
                alt.Tooltip("date:T", title="날짜"),
                alt.Tooltip("stock:Q", title="예상 재고", format=",.0f"),
                alt.Tooltip("incoming:Q", title="입고", format=",.0f"),
            ],
        )
    )

    layers = [line.mark_area(color=COLOR_LINE, opacity=0.12), line]

    if material.safety_stock > 0:
        layers.append(
            alt.Chart(pd.DataFrame({"y": [material.safety_stock]}))
            .mark_rule(color=COLOR_SAFETY, strokeDash=[5, 4], strokeWidth=1.5)
            .encode(y="y:Q")
        )

    incoming_points = frame[frame["incoming"] > 0]
    if not incoming_points.empty:
        layers.append(
            alt.Chart(incoming_points)
            .mark_point(color=COLOR_INCOMING, size=70, filled=True)
            .encode(
                x="date:T",
                y="stock:Q",
                tooltip=[
                    alt.Tooltip("date:T", title="입고일"),
                    alt.Tooltip("incoming:Q", title="입고량", format=",.0f"),
                ],
            )
        )

    # 창 밖의 소진일에 마커를 그리면 축이 그쪽까지 늘어나 창을 좁힌 의미가 없어진다.
    depletion_in_window = (
        status.days_until_depletion is not None and status.days_until_depletion <= days
    )
    if status.depletion_date and depletion_in_window:
        layers.append(
            alt.Chart(pd.DataFrame({"x": [pd.to_datetime(status.depletion_date)]}))
            .mark_rule(color=COLOR_DEPLETION, strokeDash=[3, 3], strokeWidth=1.5)
            .encode(x="x:T")
        )

    st.altair_chart(alt.layer(*layers).properties(height=280), width="stretch")

    caption = [
        f"오늘 ~ {points[-1]['date']} ({days}일)",
        "🔵 예상 재고",
        f"🟠 안전재고 {fmt_num(material.safety_stock)}{material.unit}",
    ]
    if not incoming_points.empty:
        caption.append("🟢 입고 예정")
    if status.depletion_date:
        # 창 밖이면 마커가 없으므로 그 사실을 글로 알려준다.
        caption.append(
            f"🔴 소진 {status.depletion_date}"
            if depletion_in_window
            else f"소진 {status.depletion_date} (창 밖)"
        )
    st.caption(" · ".join(caption))


def build_status_frame(material_list: list[Material], schedule_list: list[IncomingSchedule], today: str):
    """원료 + 상태를 한 번에 계산해 (rows, DataFrame)로."""
    rows = []
    for material in material_list:
        status = evaluate_material(material, schedules_for(material.id, schedule_list), today)
        rows.append((material, status))
    return rows


# ===========================================================================
# 화면 1 — 대시보드
# ===========================================================================


def page_dashboard(today: str) -> None:
    st.title("📊 대시보드")

    material_list = materials()
    schedule_list = schedules()

    if not material_list:
        st.info("등록된 원료가 없습니다. 원료를 등록하면 소진 예정일과 발주 시점을 자동으로 계산해 드립니다.")
        col1, col2 = st.columns([1, 4])
        with col1:
            if st.button("샘플 데이터 넣어보기", type="primary"):
                seed_sample_data(today)
                st.rerun()
        return

    st.caption(f"{format_korean_date(today)} 기준 수급 현황입니다.")

    rows = build_status_frame(material_list, schedule_list, today)

    overdue = [r for r in rows if r[1].alert_status == "overdue"]
    warning = [r for r in rows if r[1].alert_status == "warning"]
    depleting = [r for r in rows if is_depleting_within(r[1], 7)]

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("전체 원료", f"{len(material_list)}종")
    c2.metric("🔴 발주 지연", f"{len(overdue)}종")
    c3.metric("🟡 발주 임박", f"{len(warning)}종")
    c4.metric("⏳ 7일 내 소진", f"{len(depleting)}종")

    st.divider()

    # --- 발주 알림 ---------------------------------------------------------
    st.subheader("발주 알림")
    alert_rows = sorted(
        [r for r in rows if r[1].alert_status != "safe"], key=lambda r: sort_key_by_alert(r[1])
    )

    if not alert_rows:
        st.success("현재 조치가 필요한 원료가 없습니다.")
    else:
        st.dataframe(
            pd.DataFrame(
                [
                    {
                        "상태": status_chip(status.alert_status),
                        "원료명": material.name,
                        "현재 재고": f"{material.current_stock:,.10g} {material.unit}",
                        "발주 필요일": dday_text(status.reorder_date, today, "불필요"),
                        "소진 예정일": dday_text(status.depletion_date, today, "1년 내 없음"),
                        "공급업체": material.supplier or "-",
                    }
                    for material, status in alert_rows
                ]
            ),
            width="stretch",
            hide_index=True,
        )
        st.caption("원료를 자세히 보려면 왼쪽 메뉴의 **원료 관리**에서 선택하세요.")

    st.divider()

    # --- 7일 내 입고 예정 ---------------------------------------------------
    st.subheader("7일 내 입고 예정")
    limit = add_days(today, 7)
    name_by_id = {m.id: m for m in material_list}
    upcoming = sorted(
        [
            s
            for s in schedule_list
            if s.status == "scheduled"
            and s.material_id in name_by_id
            and diff_days(today, s.expected_date) >= 0
            and diff_days(s.expected_date, limit) >= 0
        ],
        key=lambda s: s.expected_date,
    )

    if not upcoming:
        st.info("예정된 입고가 없습니다.")
    else:
        st.dataframe(
            pd.DataFrame(
                [
                    {
                        "입고 예정일": f"{format_short_date(s.expected_date)} ({format_dday(s.expected_date, today)})",
                        "원료명": name_by_id[s.material_id].name,
                        "수량": f"+{s.quantity:,.10g} {name_by_id[s.material_id].unit}",
                        "비고": s.memo or "-",
                    }
                    for s in upcoming
                ]
            ),
            width="stretch",
            hide_index=True,
        )


# ===========================================================================
# 화면 2 — 원료 관리 (목록 + 상세)
# ===========================================================================


def page_materials(today: str) -> None:
    selected_id = st.session_state.get("selected_material_id")
    if selected_id and find_material(selected_id):
        render_material_detail(selected_id, today)
    else:
        st.session_state.selected_material_id = None
        render_material_list(today)


def render_material_list(today: str) -> None:
    st.title("📦 원료 관리")

    material_list = materials()
    schedule_list = schedules()
    st.caption(f"{len(material_list)}종 등록됨")

    with st.expander("＋ 원료 등록", expanded=not material_list):
        render_material_form(None, today)

    if not material_list:
        if st.button("샘플 데이터 넣어보기"):
            seed_sample_data(today)
            st.rerun()
        return

    rows = build_status_frame(material_list, schedule_list, today)

    # --- 검색 / 필터 / 정렬 -------------------------------------------------
    c1, c2, c3 = st.columns([2, 2, 2])
    query = c1.text_input("검색", placeholder="원료명 또는 공급업체", label_visibility="collapsed")
    status_filter = c2.selectbox(
        "상태", ["전체", "발주 지연", "발주 임박", "안전"], label_visibility="collapsed"
    )
    sort_key = c3.selectbox(
        "정렬", ["소진 임박순", "원료명순", "재고 적은순"], label_visibility="collapsed"
    )

    needle = query.strip().lower()
    visible = [
        (m, s)
        for m, s in rows
        if (status_filter == "전체" or STATUS_LABELS[s.alert_status] == status_filter)
        and (not needle or needle in m.name.lower() or needle in (m.supplier or "").lower())
    ]

    if sort_key == "원료명순":
        visible.sort(key=lambda r: r[0].name)
    elif sort_key == "재고 적은순":
        visible.sort(key=lambda r: r[0].current_stock)
    else:
        # "1년 내 소진 없음"(None)은 맨 뒤로
        visible.sort(
            key=lambda r: r[1].days_until_depletion
            if r[1].days_until_depletion is not None
            else 10**9
        )

    if not visible:
        st.info("조건에 맞는 원료가 없습니다. 검색어나 필터를 바꿔 보세요.")
        return

    st.dataframe(
        pd.DataFrame(
            [
                {
                    "상태": status_chip(status.alert_status),
                    "원료명": material.name,
                    "현재 재고": f"{material.current_stock:,.10g} {material.unit}",
                    "일사용량": f"{material.daily_usage:,.10g} {material.unit}",
                    "소진 예정일": dday_text(status.depletion_date, today, "1년 내 없음"),
                    "발주 필요일": dday_text(status.reorder_date, today, "불필요"),
                    "공급업체": material.supplier or "-",
                }
                for material, status in visible
            ]
        ),
        width="stretch",
        hide_index=True,
    )

    st.divider()
    c1, c2 = st.columns([3, 1])
    choice = c1.selectbox(
        "원료 선택",
        options=[m.id for m, _ in visible],
        format_func=lambda mid: next(m.name for m, _ in visible if m.id == mid),
    )
    c2.write("")
    if c2.button("상세 보기 →", type="primary", width="stretch"):
        st.session_state.selected_material_id = choice
        st.rerun()


def render_material_form(material: Material | None, today: str) -> None:
    """
    등록과 수정이 같은 폼을 쓴다.

    INV-4 적응: number_input은 빈 값이 될 수 없어 '빈 문자열이 0이 되는' 함정은 없지만,
    기본값 0이 그대로 저장되는 같은 성질의 함정이 남는다. daily_usage > 0 검증이 그걸 막는다.
    """
    is_edit = material is not None
    prefix = f"edit_{material.id}" if is_edit else "create"

    with st.form(f"{prefix}_material_form", clear_on_submit=not is_edit):
        c1, c2 = st.columns(2)
        name = c1.text_input("원료명 *", value=material.name if is_edit else "", max_chars=50)
        # 단위는 자유 입력이다. selectbox로 고정하면 사전에 없는 단위를 못 쓴다.
        unit = c2.text_input(
            "단위 *",
            value=material.unit if is_edit else "kg",
            max_chars=10,
            help="예: " + ", ".join(UNIT_OPTIONS),
        )

        c1, c2 = st.columns(2)
        current_stock = c1.number_input(
            "현재 재고 *",
            min_value=0.0,
            value=float(material.current_stock) if is_edit else 0.0,
            step=1.0,
            format="%.2f",
        )
        daily_usage = c2.number_input(
            "일일 평균 사용량 *",
            min_value=0.0,
            value=float(material.daily_usage) if is_edit else 0.0,
            step=1.0,
            format="%.2f",
            help="0보다 커야 소진일 계산이 가능합니다.",
        )

        c1, c2, c3 = st.columns(3)
        safety_stock = c1.number_input(
            "안전재고 *",
            min_value=0.0,
            value=float(material.safety_stock) if is_edit else 0.0,
            step=1.0,
            format="%.2f",
            help="이 아래로 내려가면 안 되는 최소 재고",
        )
        lead_time_days = c2.number_input(
            "리드타임(일) *",
            min_value=0,
            value=int(material.lead_time_days) if is_edit else 7,
            step=1,
            help="발주 후 입고까지 걸리는 일수",
        )
        alert_lead_days = c3.number_input(
            "알림 선행일 *",
            min_value=0,
            value=int(material.alert_lead_days) if is_edit else 3,
            step=1,
            help="발주 필요일 며칠 전부터 '발주 임박'으로 표시할지",
        )

        supplier = st.text_input("공급업체", value=material.supplier if is_edit else "")
        memo = st.text_area(
            "메모", value=material.memo if is_edit else "", max_chars=500, height=80
        )

        submitted = st.form_submit_button("저장" if is_edit else "등록", type="primary")

    if not submitted:
        return

    existing_names = [m.name for m in materials() if not is_edit or m.id != material.id]
    errors = validate_material(
        name, unit, current_stock, daily_usage, lead_time_days, safety_stock,
        alert_lead_days, memo, existing_names,
    )

    if errors:
        for message in errors:
            st.error(message)
        return

    payload = dict(
        name=name.strip(),
        unit=unit.strip(),
        current_stock=float(current_stock),
        daily_usage=float(daily_usage),
        lead_time_days=int(lead_time_days),
        safety_stock=float(safety_stock),
        alert_lead_days=int(alert_lead_days),
        supplier=supplier.strip(),
        memo=memo.strip(),
    )

    try:
        if is_edit:
            update_material(material.id, **payload)
            st.toast("저장했습니다.", icon="✅")
        else:
            created = create_material(**payload)
            st.session_state.selected_material_id = created.id
            st.toast(f"'{created.name}'을(를) 등록했습니다.", icon="✅")
        st.rerun()
    except StorageError as error:
        st.error(str(error))


def render_material_detail(material_id: str, today: str) -> None:
    material = find_material(material_id)
    if material is None:
        st.session_state.selected_material_id = None
        st.rerun()
        return

    own_schedules = schedules_for(material.id)
    status = evaluate_material(material, own_schedules, today)

    c1, c2 = st.columns([4, 1])
    c1.title(f"{material.name} {status_chip(status.alert_status)}")
    c2.write("")
    if c2.button("← 목록으로", width="stretch"):
        st.session_state.selected_material_id = None
        st.rerun()

    st.caption(
        f"마지막 수정 {material.updated_at[:10]}"
        + (f" · {material.supplier}" if material.supplier else "")
    )

    # --- 상태 패널 ---------------------------------------------------------
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("현재 재고", f"{material.current_stock:,.10g} {material.unit}")
    c2.metric(
        "소진 예정일 (D-Day)",
        format_dday(status.depletion_date, today) if status.depletion_date else "없음",
        help=format_korean_date(status.depletion_date) if status.depletion_date else "1년 내 소진 없음",
    )
    c3.metric(
        "발주 필요일",
        format_dday(status.reorder_date, today) if status.reorder_date else "불필요",
        help=format_korean_date(status.reorder_date) if status.reorder_date else "1년 내 발주 불필요",
    )
    c4.metric(
        "발주 임계값",
        f"{status.reorder_threshold:,.10g} {material.unit}",
        help=f"안전재고 {fmt_num(material.safety_stock)} + 리드타임 {material.lead_time_days}일분",
    )

    if status.is_overdue:
        st.error(
            "재고가 발주 임계값 이하입니다. **지금 발주해야** 리드타임 안에 안전재고를 지킬 수 있습니다."
        )
    elif status.alert_status == "warning":
        st.warning(
            f"발주 필요일이 {status.days_until_reorder}일 남았습니다. 발주 준비를 시작하세요."
        )

    # 기본은 데이터에 맞춘 자동 창. 더 멀리 보고 싶으면 직접 고를 수 있게 해 둔다.
    auto_days = choose_projection_days(material, own_schedules, status, today)
    RANGE_OPTIONS: dict[str, int | None] = {
        f"자동 ({auto_days}일)": None,
        "30일": 30,
        "60일": 60,
        "90일": 90,
        "180일": 180,
    }

    c1, c2 = st.columns([3, 1])
    c1.subheader("재고 추이 예측")
    choice = c2.selectbox(
        "표시 기간",
        list(RANGE_OPTIONS.keys()),
        # 원료마다 자동값이 다르므로 키에 id를 넣어 위젯이 섞이지 않게 한다.
        key=f"range_{material.id}",
        label_visibility="collapsed",
    )
    stock_chart(material, own_schedules, status, today, RANGE_OPTIONS[choice] or auto_days)

    tab_info, tab_schedule, tab_history = st.tabs(["기본 정보", "입고 예정", "재고 변동 이력"])

    # --- 기본 정보 ---------------------------------------------------------
    with tab_info:
        render_material_form(material, today)

        st.divider()
        with st.expander("⚠️ 원료 삭제"):
            st.write(
                f"**{material.name}**을(를) 삭제합니다. 연결된 입고 예정과 재고 변동 이력도 "
                "함께 삭제되며 되돌릴 수 없습니다."
            )
            confirm = st.text_input("확인을 위해 원료명을 그대로 입력하세요", key="delete_confirm")
            if st.button("삭제", type="primary", disabled=confirm != material.name):
                delete_material(material.id)
                st.session_state.selected_material_id = None
                st.toast("삭제했습니다.", icon="🗑")
                st.rerun()

    # --- 입고 예정 ---------------------------------------------------------
    with tab_schedule:
        # st.rerun()은 폼 컨테이너 밖에서 호출한다.
        with st.form("add_schedule_form", clear_on_submit=True):
            c1, c2, c3 = st.columns([2, 2, 3])
            expected = c1.date_input("입고 예정일 *", value=parse_date(add_days(today, 7)))
            quantity = c2.number_input("수량 *", min_value=0.0, value=0.0, step=1.0, format="%.2f")
            memo = c3.text_input("비고", max_chars=200, placeholder="발주번호, 담당자 등")
            schedule_submitted = st.form_submit_button("입고 예정 추가", type="primary")

        if schedule_submitted:
            errors = validate_schedule(expected.isoformat(), quantity, memo)
            if errors:
                for message in errors:
                    st.error(message)
            else:
                create_schedule(material.id, expected.isoformat(), float(quantity), memo)
                st.toast("입고 예정을 추가했습니다.", icon="✅")
                st.rerun()

        if not own_schedules:
            st.info("등록된 입고 예정이 없습니다.")
        else:
            for schedule in sorted(own_schedules, key=lambda s: s.expected_date):
                is_late = (
                    schedule.status == "scheduled" and diff_days(today, schedule.expected_date) < 0
                )
                label = SCHEDULE_STATUS_LABELS[schedule.status]
                if is_late:
                    label = "⚠️ 입고 지연"
                elif schedule.status == "completed":
                    label = "✅ 입고 완료"
                elif schedule.status == "cancelled":
                    label = "✖ 취소됨"

                c1, c2, c3, c4 = st.columns([2, 2, 2, 3])
                c1.write(f"**{format_short_date(schedule.expected_date)}**")
                c2.write(f"+{schedule.quantity:,.10g} {material.unit}")
                c3.write(label)

                with c4:
                    if schedule.status == "scheduled":
                        b1, b2 = st.columns(2)
                        if b1.button("입고 완료", key=f"complete_{schedule.id}", type="primary"):
                            if complete_schedule(schedule.id, today):
                                st.toast(
                                    f"입고 {schedule.quantity:,.10g}{material.unit}을(를) 재고에 반영했습니다.",
                                    icon="📦",
                                )
                            st.rerun()
                        if b2.button("취소", key=f"cancel_{schedule.id}"):
                            cancel_schedule(schedule.id)
                            st.toast("입고 예정을 취소했습니다.", icon="✖")
                            st.rerun()
                    else:
                        if st.button("삭제", key=f"delete_sched_{schedule.id}"):
                            delete_schedule(schedule.id)
                            st.rerun()

                if schedule.memo:
                    st.caption(f"　비고: {schedule.memo}")

    # --- 변동 이력 ---------------------------------------------------------
    with tab_history:
        with st.form("adjust_form", clear_on_submit=True):
            st.write(f"현재 재고: **{material.current_stock:,.10g} {material.unit}**")
            c1, c2 = st.columns([2, 3])
            new_stock = c1.number_input(
                "조정 후 재고 *",
                min_value=0.0,
                value=float(material.current_stock),
                step=1.0,
                format="%.2f",
            )
            reason = c2.text_input("조정 사유 *", placeholder="예: 실사 차이, 폐기, 반품")
            adjust_submitted = st.form_submit_button("재고 조정", type="primary")

        if adjust_submitted:
            if not reason.strip():
                st.error("조정 사유를 입력하세요.")
            elif new_stock == material.current_stock:
                st.warning("재고 변동이 없습니다.")
            else:
                adjust_stock(material.id, float(new_stock), reason, today)
                st.toast("재고를 조정했습니다.", icon="✅")
                st.rerun()

        material_logs = sorted(
            [l for l in logs() if l.material_id == material.id],
            key=lambda l: l.created_at,
            reverse=True,
        )

        if not material_logs:
            st.info("변동 이력이 없습니다.")
        else:
            st.dataframe(
                pd.DataFrame(
                    [
                        {
                            "날짜": log.date,
                            "유형": LOG_TYPE_LABELS[log.type],
                            "변동": f"{'+' if log.delta >= 0 else ''}{log.delta:,.10g} {material.unit}",
                            "변동 후": f"{log.stock_after:,.10g} {material.unit}",
                            "사유": log.reason or "-",
                        }
                        for log in material_logs
                    ]
                ),
                width="stretch",
                hide_index=True,
            )


# ===========================================================================
# 화면 3 — 주간 리포트
# ===========================================================================


def page_report(today: str) -> None:
    st.title("📄 주간 수급 리포트")

    material_list = materials()
    schedule_list = schedules()

    if "week_start" not in st.session_state:
        st.session_state.week_start = start_of_week(today)

    c1, c2, c3, c4 = st.columns([1, 2, 1, 1])
    if c1.button("‹ 이전 주", width="stretch"):
        st.session_state.week_start = add_days(st.session_state.week_start, -7)
        st.rerun()

    picked = c2.date_input(
        "기준 주 시작(월)",
        value=parse_date(st.session_state.week_start),
        label_visibility="collapsed",
    )
    # 사용자가 주 중간 날짜를 골라도 그 주의 월요일로 정규화한다.
    normalized = start_of_week(picked.isoformat())
    if normalized != st.session_state.week_start:
        st.session_state.week_start = normalized
        st.rerun()

    if c3.button("다음 주 ›", width="stretch"):
        st.session_state.week_start = add_days(st.session_state.week_start, 7)
        st.rerun()
    if c4.button("이번 주", width="stretch"):
        st.session_state.week_start = start_of_week(today)
        st.rerun()

    week_start = st.session_state.week_start
    week_end = end_of_week(week_start)
    st.caption(f"{format_korean_date(week_start)} ~ {format_korean_date(week_end)}")

    risky_only = st.checkbox("위험 원료만 (발주 지연·임박)")

    # 화면 표 / TSV / 엑셀이 전부 이 결과 하나를 쓴다. 두 벌로 만들면 반드시 어긋난다.
    rows = build_report_rows(
        material_list, schedule_list, week_start, week_end, today, risky_only
    )

    counts = {"overdue": 0, "warning": 0, "safe": 0}
    for row in rows:
        counts[row.status.alert_status] += 1

    c1, c2, c3 = st.columns(3)
    c1.metric("🔴 발주 지연", f"{counts['overdue']}종")
    c2.metric("🟡 발주 임박", f"{counts['warning']}종")
    c3.metric("🟢 안전", f"{counts['safe']}종")

    if not rows:
        st.info(
            "등록된 원료가 없습니다." if not material_list else "조건에 맞는 원료가 없습니다."
        )
        return

    st.dataframe(
        pd.DataFrame([row.cells for row in rows], columns=REPORT_HEADERS),
        width="stretch",
        hide_index=True,
    )

    st.divider()
    st.subheader("내보내기")

    c1, c2 = st.columns(2)
    with c1:
        st.download_button(
            "⬇ 엑셀 다운로드 (.xlsx)",
            data=rows_to_excel(rows, week_start, week_end),
            file_name=f"원료수급리포트_{week_start}_{week_end}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            width="stretch",
        )
    with c2:
        st.download_button(
            "⬇ CSV 다운로드",
            # BOM을 붙여야 엑셀에서 한글이 깨지지 않는다.
            data=("﻿" + rows_to_csv(rows)).encode("utf-8"),
            file_name=f"원료수급리포트_{week_start}_{week_end}.csv",
            mime="text/csv",
            width="stretch",
        )

    with st.expander("📋 표 복사 (TSV — 엑셀·메일에 그대로 붙여넣기)"):
        # st.code 오른쪽 위의 복사 버튼을 그대로 쓴다.
        st.code(rows_to_tsv(rows), language=None)

    st.caption("표·엑셀·복사는 모두 화면에 보이는 것과 **같은 데이터**를 사용합니다. (필터 포함)")


# ===========================================================================
# 화면 4 — AI 어시스턴트
# ===========================================================================

SUGGESTIONS = [
    "이번 주에 발주해야 할 원료 알려줘",
    "가장 위험한 원료 3개는?",
    "다음 주 입고 예정 정리해줘",
    "재고가 안전재고 아래인 원료 있어?",
]


def ensure_index(today: str, force: bool = False) -> VectorIndex | None:
    """
    인덱스가 없거나 데이터가 바뀌었으면 다시 만든다.
    안 그러면 옛날 재고로 답한다.
    """
    version = st.session_state.get("data_version", 0)
    index: VectorIndex | None = st.session_state.get("rag_index")

    if index is not None and index.data_version == version and not force:
        return index

    chunks = build_chunks(materials(), schedules(), logs(), today)
    if not chunks:
        st.session_state.rag_index = None
        return None

    with st.spinner("인덱스를 만드는 중… (변경된 문장만 임베딩합니다)"):
        new_index, embedded, reused = build_index(chunks)

    new_index.data_version = version
    st.session_state.rag_index = new_index
    st.session_state.last_index_stats = (len(chunks), embedded, reused)
    return new_index


def page_assistant(today: str) -> None:
    st.title("💬 AI 어시스턴트")
    st.caption("재고·입고·수급에 대해 물어보세요. 관련 문장만 골라 답합니다.")

    api_key = get_api_key()

    if not api_key:
        st.error(
            "**OPENAI_API_KEY가 설정되지 않았습니다.**\n\n"
            "프로젝트 루트의 `.env`에 키를 넣고 앱을 다시 시작해 주세요.\n\n"
            "재고 관리·소진 예측·발주 알림·주간 리포트는 키 없이도 정상 동작합니다."
        )
        return

    if not materials():
        st.info("등록된 원료가 없습니다. 원료를 먼저 등록하면 질문할 수 있습니다.")
        return

    c1, c2 = st.columns([3, 1])
    c1.caption(f"모델 `{get_chat_model()}` · 임베딩 `{get_embedding_model()}`")
    if c2.button("🔄 인덱스 재생성", width="stretch"):
        try:
            ensure_index(today, force=True)
            total, embedded, reused = st.session_state.get("last_index_stats", (0, 0, 0))
            st.toast(
                f"인덱스 갱신 완료 — 청크 {total}개 (신규 {embedded}, 재사용 {reused})", icon="🔄"
            )
        except LlmError as error:
            st.error(str(error))

    if "chat_messages" not in st.session_state:
        st.session_state.chat_messages = []

    if not st.session_state.chat_messages:
        st.info(
            "질문하면 FAISS로 관련된 문장만 추려 답합니다. "
            "재고 데이터는 질문할 때만 OpenAI로 전송됩니다."
        )
        cols = st.columns(2)
        for i, suggestion in enumerate(SUGGESTIONS):
            if cols[i % 2].button(suggestion, key=f"suggest_{i}", width="stretch"):
                st.session_state.pending_question = suggestion
                st.rerun()

    for message in st.session_state.chat_messages:
        with st.chat_message(message["role"]):
            st.markdown(message["content"])
            if message.get("sources"):
                with st.expander(f"근거 {len(message['sources'])}건"):
                    for source in message["sources"]:
                        st.caption(
                            f"**[{source['id']}]** {source['label']} (유사도 {source['score']:.2f})"
                        )
            if message.get("prompt_tokens"):
                st.caption(f"프롬프트 {message['prompt_tokens']:,} 토큰")

    question = st.chat_input("예: 이번 주에 발주해야 할 원료 알려줘")
    if not question and "pending_question" in st.session_state:
        # 추천 질문 칩을 눌러 들어온 경우
        question = st.session_state["pending_question"]
        del st.session_state["pending_question"]
    if not question:
        return

    st.session_state.chat_messages.append({"role": "user", "content": question})

    history = [
        {"role": m["role"], "content": m["content"]}
        for m in st.session_state.chat_messages[-5:-1]
    ]

    try:
        index = ensure_index(today)
        if index is None:
            raise LlmError("인덱싱할 데이터가 없습니다.")

        with st.spinner("관련 문장을 찾는 중…"):
            result = answer_question(question, history, today, index)

        st.session_state.chat_messages.append(
            {
                "role": "assistant",
                "content": result["answer"],
                "sources": result["sources"],
                "prompt_tokens": result["prompt_tokens"],
            }
        )
    except LlmError as error:
        st.session_state.chat_messages.append(
            {"role": "assistant", "content": f"⚠️ {error}"}
        )
    except Exception as error:  # noqa: BLE001 - AI 실패가 앱을 죽이면 안 된다
        logger.exception("질의응답 실패")
        st.session_state.chat_messages.append(
            {"role": "assistant", "content": f"⚠️ 예기치 못한 오류: {error}"}
        )

    st.rerun()


# ===========================================================================
# 엔트리포인트
# ===========================================================================

PAGES = {
    "📊 대시보드": page_dashboard,
    "📦 원료 관리": page_materials,
    "📄 주간 리포트": page_report,
    "💬 AI 어시스턴트": page_assistant,
}


def main() -> None:
    st.set_page_config(
        page_title="원료 재고 소진 예측 & 발주 알림",
        page_icon="🧪",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    # "오늘"은 한 번만 정해 화면 전체가 같은 기준일을 쓰게 한다.
    # (자정을 넘겨 열어둔 세션은 새로고침하면 갱신된다)
    today = today_str()

    with st.sidebar:
        st.title("🧪 원료 재고 예측")
        st.caption("소진 예측 & 발주 알림")

        page_name = st.radio("메뉴", list(PAGES.keys()), label_visibility="collapsed")

        st.divider()
        st.caption(format_korean_date(today))

        material_list = materials()
        if material_list:
            rows = build_status_frame(material_list, schedules(), today)
            overdue = sum(1 for _, s in rows if s.alert_status == "overdue")
            warning = sum(1 for _, s in rows if s.alert_status == "warning")
            if overdue:
                st.error(f"발주 지연 {overdue}종")
            if warning:
                st.warning(f"발주 임박 {warning}종")
            if not overdue and not warning:
                st.success("모든 원료 안전")

        with st.expander("데이터 관리"):
            st.caption(f"저장 위치: `{STORE_PATH.relative_to(APP_ROOT)}`")
            st.download_button(
                "백업 내려받기 (JSON)",
                data=json.dumps(get_store(), ensure_ascii=False, indent=2).encode("utf-8"),
                file_name=f"재고백업_{today}.json",
                mime="application/json",
                width="stretch",
            )
            uploaded = st.file_uploader("백업 복원", type="json", label_visibility="collapsed")
            if uploaded is not None and st.button("복원 실행", width="stretch"):
                try:
                    restored = json.loads(uploaded.getvalue().decode("utf-8"))
                    commit(
                        {
                            "materials": list(restored.get("materials", [])),
                            "schedules": list(restored.get("schedules", [])),
                            "logs": list(restored.get("logs", [])),
                        }
                    )
                    st.session_state.selected_material_id = None
                    st.toast("복원했습니다.", icon="✅")
                    st.rerun()
                except Exception as error:  # noqa: BLE001
                    st.error(f"복원 실패: {error}")

    # 원료 관리 메뉴를 벗어나면 상세 선택을 해제한다.
    if page_name != "📦 원료 관리":
        st.session_state.selected_material_id = None

    PAGES[page_name](today)


if __name__ == "__main__":
    main()
